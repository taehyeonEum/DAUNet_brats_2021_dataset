import numpy as np
import openpyxl as xl
import os
import shutil

# image = np.array(nib.load(path).get_fdata())

old_train_img_path = '/home/NAS_mount/thum/brats/data/Train_img'
old_train_msk_path = '/home/NAS_mount/thum/brats/data/Train_msk'

old_valid_img_path = '/home/NAS_mount/thum/brats/data/Valid_img'
old_valid_msk_path = '/home/NAS_mount/thum/brats/data/Valid_msk'

new_train_img_path = '/home/NAS_mount/thum/brats/data_T/Train_img'
new_train_msk_path = '/home/NAS_mount/thum/brats/data_T/Train_msk'

new_valid_img_path = '/home/NAS_mount/thum/brats/data_T/Valid_img'
new_valid_msk_path = '/home/NAS_mount/thum/brats/data_T/Valid_msk'

xl_path = '/home/NAS_mount/thum/brats/Annotation_brats_thum.xlsx'



if not os.path.exists(new_train_img_path): os.mkdir(new_train_img_path)
if not os.path.exists(new_train_msk_path): os.mkdir(new_train_msk_path)
if not os.path.exists(new_valid_img_path): os.mkdir(new_valid_img_path)
if not os.path.exists(new_valid_msk_path): os.mkdir(new_valid_msk_path)

old_train_imgs = os.listdir(old_train_img_path)

#make tumor index tuple list
wb = xl.load_workbook(xl_path)
ws = wb.active

tumor_slices = []
for i in range(900):
    x = i+2
    ts = int(ws.cell(row = x, column=4).value)
    te = int(ws.cell(row = x, column=5).value)

    tumor_slices.append(((ts + i*155),(te + i*155)))

print(tumor_slices)

ts_idx = 0
for idx, img in enumerate(old_train_imgs):

    if idx < tumor_slices[ts_idx][0]: continue

    if idx >= tumor_slices[ts_idx][0] & idx < tumor_slices[ts_idx][1]:
        shutil.copy('______', '_____________')

    img_path_1 = os.path.join(old_train_img_path, img)



