import numpy as np
import openpyxl as xl
import os
import nibabel as nb
import matplotlib.pyplot as plt
from skimage.transform import resize

# image = np.array(nib.load(path).get_fdata())

base_dir = '/home/NAS_mount/thum/_raw_data/BraTS2021_Training_Data'
Anno_xl_path = '/home/NAS_mount/thum/brats/Annotation_brats_thum.xlsx'

train_img_dir = '/home/NAS_mount/thum/brats/data2/Train_img'
if not os.path.exists(train_img_dir): os.mkdir(train_img_dir)
train_msk_dir = '/home/NAS_mount/thum/brats/data2/Train_msk'
if not os.path.exists(train_msk_dir): os.mkdir(train_msk_dir)
valid_img_dir = '/home/NAS_mount/thum/brats/data2/Valid_img'
if not os.path.exists(valid_img_dir): os.mkdir(valid_img_dir)
valid_msk_dir = '/home/NAS_mount/thum/brats/data2/Valid_msk'
if not os.path.exists(valid_msk_dir): os.mkdir(valid_msk_dir)
test_img_dir = '/home/NAS_mount/thum/brats/data2/Test_img'
if not os.path.exists(test_img_dir): os.mkdir(test_img_dir)
test_msk_dir = '/home/NAS_mount/thum/brats/data2/Test_msk'
if not os.path.exists(test_msk_dir): os.mkdir(test_msk_dir)

patient_list = os.listdir(base_dir)
wb = xl.load_workbook(Anno_xl_path)
ws = wb.active

def str_num(num):
    if num < 10:
        return '00'+str(num)
    elif num < 100 :
        return '0' + str(num)
    else: return str(num)


for idx, sub_name in enumerate(patient_list): # sub_name is subject name

    print('_______________{}_______________'.format(sub_name))

    tTop = ws.cell(row=(idx+2), column=4).value
    tBottom = ws.cell(row=(idx+2), column=5).value

    sub_path = os.path.join(base_dir, sub_name)

    contrast_names = os.listdir(sub_path)

    # param = {}
    conts = {}

    for cont_name in contrast_names:
        cont_path = os.path.join(sub_path, cont_name) 

        cont_type = cont_name.split('.')[0].split('_')[-1] # type of contrast

        if cont_type == 'seg':
            cont_tr = np.array(nb.load(cont_path).get_fdata()) # npy array

            cont_tr = cont_tr.transpose(2, 0, 1)
        # print(cont_tr.shape) # 155, 240, 240

        # cont_tr = cont_tr/cont_tr.max()

        # if cont_type == 'flair':
        
        # param['{}_mean'.format(cont_type)] = cont_tr.mean()
        # param['max'] = cont_tr.max()
        # param['min'] = cont_tr.min()
        # param['{}_std'.format(cont_type)] = cont_tr.std()

        '''important lines'''
        # conts[cont_type] = (cont_tr - cont_tr.min()) / (cont_tr.max() - cont_tr.min())

        # print('mean', cont_tr.mean())
        # print('max', cont_tr.max())
        # print('min', cont_tr.min())
        # print('std', cont_tr.std())        

        # for i in range(155):
            

        #     print('--------{}_{}_{}--------'.format(sub_name, cont_type, i))

            # original = cont_tr[i]
            # print('original sum ', original.sum())
            # print('original mean ',original.mean())
            # print('origianal.max() - origianl.min()', original.max() - original.min() )

            # min_max = (cont_tr[i] - param['min'])/(param['max'] - param['min'])
            # print('min max sum ', min_max.sum())
            # print('min max mean ',min_max.mean())
            # print('min max.max() - min max.min()', min_max.max() - min_max.min() )

            
            

            # plt.rcParams["figure.figsize"] = (15,5)
            # plt.subplot(1, 3, 1)
            # plt.imshow(original)
            # plt.subplot(1, 3, 2)
            # plt.title(cont_type + str(i))
            # plt.imshow(min_max)
            # plt.subplot(1, 3, 3)
            # plt.imshow(z_score)
            # plt.show()
    
    # print(conts.keys())
    # print(conts['seg'].shape)

    # concated = np.concatenate((np.expand_dims(conts['t1'], 1), np.expand_dims(conts['t1ce'], 1), np.expand_dims(conts['t2'], 1), np.expand_dims(conts['flair'], 1) ), 1)
    # print(concated.shape)



    for idx_s, slice in enumerate(cont_tr):

        tumor = 'N'
        if idx_s >= tTop and idx_s <= tBottom:
            tumor = 'T'

        if idx < 900:
            save_dir = (train_img_dir, train_msk_dir)
        elif idx < 1050:
            save_dir = (valid_img_dir, valid_msk_dir)
        else:
            save_dir = (test_img_dir, test_msk_dir)

        msk_path = os.path.join( save_dir[1] , sub_name + '_' + str_num(idx_s) + '_{}.npy'.format(tumor))

        # for i in range(155):
        np.save(msk_path, (cont_tr[idx_s]))


    # if idx == 1: break